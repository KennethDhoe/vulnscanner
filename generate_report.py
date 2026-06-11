#!/usr/bin/env python3
"""
generate_report.py — Merge Syft + Grype scan output into an Excel report
Usage: python generate_report.py [scan_output_dir] [report.xlsx]
"""

import json
import sys
import os
import glob
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# ── Config ───────────────────────────────────────────────────────────────────
SCAN_DIR = sys.argv[1] if len(sys.argv) > 1 else "./scan_output"
OUTPUT_FILE = sys.argv[2] if len(sys.argv) > 2 else "security_report.xlsx"

SEVERITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Negligible": 4, "Unknown": 5}

SEVERITY_COLORS = {
    "Critical": "C00000",
    "High":     "FF0000",
    "Medium":   "FF7C00",
    "Low":      "FFD700",
    "Negligible":"A9A9A9",
    "Unknown":  "D3D3D3",
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def header_font(bold=True, color="FFFFFF", size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)

def cell_font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header_row(ws, headers, row=1, bg="1F4E79", fg="FFFFFF"):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font(color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 28

def write_data_row(ws, row_num, values, severity=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = cell_font()
        c.alignment = Alignment(vertical="center", wrap_text=False)
        c.border = thin_border()
        if row_num % 2 == 0:
            c.fill = fill("F5F7FA")
    if severity and severity in SEVERITY_COLORS:
        sev_col = None
        for col, val in enumerate(values, 1):
            if val == severity:
                sev_col = col
                break
        if sev_col:
            c = ws.cell(row=row_num, column=sev_col)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill = fill(SEVERITY_COLORS[severity])

def auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

def freeze_and_filter(ws, freeze="A2", filter_range=None):
    ws.freeze_panes = freeze
    if filter_range:
        ws.auto_filter.ref = filter_range

# ── Data loaders ──────────────────────────────────────────────────────────────
def load_json(path):
    if not os.path.exists(path):
        return None
    with open(path) as f:
        return json.load(f)

def load_host_meta(path):
    meta = {}
    if not os.path.exists(path):
        return meta
    with open(path) as f:
        for line in f:
            k, _, v = line.strip().partition("=")
            meta[k] = v
    return meta

def parse_syft_packages(sbom):
    if not sbom or "artifacts" not in sbom:
        return []
    packages = []
    for a in sbom["artifacts"]:
        packages.append({
            "name":     a.get("name", ""),
            "version":  a.get("version", ""),
            "type":     a.get("type", ""),
            "language": a.get("language", ""),
            "location": "; ".join(
                loc.get("realPath", loc.get("path", ""))
                for loc in a.get("locations", [])
            )[:200],
            "licenses": ", ".join(
                (lic.get("value") or lic) if isinstance(lic, dict) else str(lic)
                for lic in a.get("licenses", [])
            ),
        })
    return packages

def parse_grype_vulns(vuln_json, source="host"):
    if not vuln_json or "matches" not in vuln_json:
        return []
    vulns = []
    for m in vuln_json["matches"]:
        v = m.get("vulnerability", {})
        art = m.get("artifact", {})
        fix = v.get("fix", {})
        fix_versions = ", ".join(fix.get("versions", [])) if fix else ""
        fix_state = fix.get("state", "") if fix else ""
        urls = "; ".join(v.get("urls", []))[:300]
        vulns.append({
            "source":        source,
            "cve":           v.get("id", ""),
            "severity":      v.get("severity", "Unknown"),
            "cvss":          next(
                (s.get("metrics", {}).get("baseScore", "")
                 for s in v.get("cvss", []) if s.get("metrics", {}).get("baseScore")),
                ""
            ),
            "package":       art.get("name", ""),
            "version":       art.get("version", ""),
            "pkg_type":      art.get("type", ""),
            "fix_version":   fix_versions,
            "fix_state":     fix_state,
            "description":   v.get("description", "")[:300],
            "urls":          urls,
        })
    return vulns

# ── Sheet builders ────────────────────────────────────────────────────────────
def build_summary(ws, meta, host_pkgs, host_vulns, container_data):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    title_cell = ws["A1"]
    title_cell.value = "Security Scan Report"
    title_cell.font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 32

    rows = [
        ("Scan Date",       meta.get("date", datetime.utcnow().isoformat())),
        ("Hostname",        meta.get("hostname", "N/A")),
        ("OS",              meta.get("os", "N/A")),
        ("Kernel",          meta.get("kernel", "N/A")),
        ("Architecture",    meta.get("arch", "N/A")),
        ("Uptime",          meta.get("uptime", "N/A")),
        ("", ""),
        ("Host Packages",   len(host_pkgs)),
        ("Host Vulns",      len(host_vulns)),
        ("Critical",        sum(1 for v in host_vulns if v["severity"] == "Critical")),
        ("High",            sum(1 for v in host_vulns if v["severity"] == "High")),
        ("Medium",          sum(1 for v in host_vulns if v["severity"] == "Medium")),
        ("Low",             sum(1 for v in host_vulns if v["severity"] == "Low")),
        ("", ""),
        ("Containers Scanned", len(container_data)),
    ]

    all_container_vulns = [v for _, _, cvulns in container_data for v in cvulns]
    rows += [
        ("Container Vulns Total", len(all_container_vulns)),
        ("Container Critical",    sum(1 for v in all_container_vulns if v["severity"] == "Critical")),
        ("Container High",        sum(1 for v in all_container_vulns if v["severity"] == "High")),
    ]

    for i, (k, v) in enumerate(rows, 2):
        ws.row_dimensions[i].height = 20
        ka = ws.cell(row=i, column=1, value=k)
        va = ws.cell(row=i, column=2, value=v)
        if k:
            ka.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
            va.font = Font(name="Arial", size=10)
        if k in ("Critical", "Container Critical"):
            va.font = Font(name="Arial", bold=True, size=10, color="C00000")
        if k in ("High", "Container High"):
            va.font = Font(name="Arial", bold=True, size=10, color="FF0000")


def build_host_packages(ws, packages):
    ws.sheet_view.showGridLines = False
    headers = ["Name", "Version", "Type", "Language", "Licenses", "Location"]
    write_header_row(ws, headers)
    for i, p in enumerate(sorted(packages, key=lambda x: x["name"]), 2):
        write_data_row(ws, i, [
            p["name"], p["version"], p["type"],
            p["language"], p["licenses"], p["location"]
        ])
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    auto_width(ws)


def build_vulns_sheet(ws, vulns, label="host"):
    ws.sheet_view.showGridLines = False
    if not vulns:
        ws["A1"] = "No vulnerabilities found."
        return
    headers = ["Source", "CVE / ID", "Severity", "CVSS", "Package", "Version",
               "Type", "Fix Version", "Fix State", "Description", "References"]
    write_header_row(ws, headers)
    sorted_vulns = sorted(vulns, key=lambda v: (
        SEVERITY_ORDER.get(v["severity"], 99), v["package"]
    ))
    for i, v in enumerate(sorted_vulns, 2):
        write_data_row(ws, i, [
            v["source"], v["cve"], v["severity"], v["cvss"],
            v["package"], v["version"], v["pkg_type"],
            v["fix_version"], v["fix_state"], v["description"], v["urls"]
        ], severity=v["severity"])
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    auto_width(ws)


def build_container_inventory(ws, container_data):
    ws.sheet_view.showGridLines = False
    headers = ["Image", "Packages", "Total Vulns", "Critical", "High", "Medium", "Low"]
    write_header_row(ws, headers)
    for i, (img, pkgs, vulns) in enumerate(sorted(container_data, key=lambda x: x[0]), 2):
        crit = sum(1 for v in vulns if v["severity"] == "Critical")
        high = sum(1 for v in vulns if v["severity"] == "High")
        med  = sum(1 for v in vulns if v["severity"] == "Medium")
        low  = sum(1 for v in vulns if v["severity"] == "Low")
        write_data_row(ws, i, [img, len(pkgs), len(vulns), crit, high, med, low])
        if crit:
            ws.cell(row=i, column=4).font = Font(name="Arial", bold=True, color="C00000", size=10)
        if high:
            ws.cell(row=i, column=5).font = Font(name="Arial", bold=True, color="FF0000", size=10)
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    auto_width(ws)


def build_critical_high(ws, all_vulns):
    ws.sheet_view.showGridLines = False
    filtered = [v for v in all_vulns if v["severity"] in ("Critical", "High")]
    if not filtered:
        ws["A1"] = "No Critical or High vulnerabilities found."
        return
    headers = ["Source", "CVE / ID", "Severity", "CVSS", "Package", "Version",
               "Fix Version", "Fix State", "Description"]
    write_header_row(ws, headers, bg="7B0000")
    sorted_vulns = sorted(filtered, key=lambda v: (
        SEVERITY_ORDER.get(v["severity"], 99), -(float(v["cvss"]) if v["cvss"] else 0)
    ))
    for i, v in enumerate(sorted_vulns, 2):
        write_data_row(ws, i, [
            v["source"], v["cve"], v["severity"], v["cvss"],
            v["package"], v["version"], v["fix_version"],
            v["fix_state"], v["description"]
        ], severity=v["severity"])
    freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    auto_width(ws)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"[*] Loading scan data from: {SCAN_DIR}")

    meta        = load_host_meta(os.path.join(SCAN_DIR, "host_meta.txt"))
    host_sbom   = load_json(os.path.join(SCAN_DIR, "host_sbom.json"))
    host_vj     = load_json(os.path.join(SCAN_DIR, "host_vulns.json"))
    host_pkgs   = parse_syft_packages(host_sbom)
    host_vulns  = parse_grype_vulns(host_vj, source="host")

    print(f"    Host packages : {len(host_pkgs)}")
    print(f"    Host vulns    : {len(host_vulns)}")

    container_data = []
    containers_dir = os.path.join(SCAN_DIR, "containers")
    if os.path.isdir(containers_dir):
        for sbom_path in glob.glob(os.path.join(containers_dir, "*_sbom.json")):
            base = sbom_path.replace("_sbom.json", "")
            img_name = os.path.basename(base).replace("___", "/", 1).replace("___", ":")
            vuln_path = base + "_vulns.json"
            sbom = load_json(sbom_path)
            vj   = load_json(vuln_path)
            pkgs  = parse_syft_packages(sbom)
            vulns = parse_grype_vulns(vj, source=img_name)
            container_data.append((img_name, pkgs, vulns))
            print(f"    Container [{img_name}]: {len(pkgs)} pkgs, {len(vulns)} vulns")

    all_vulns = host_vulns + [v for _, _, cvulns in container_data for v in cvulns]

    wb = Workbook()

    # Sheet 1 — Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    build_summary(ws_summary, meta, host_pkgs, host_vulns, container_data)

    # Sheet 2 — Critical & High (priority view)
    ws_crit = wb.create_sheet("⚠ Critical & High")
    build_critical_high(ws_crit, all_vulns)

    # Sheet 3 — Host Vulnerabilities
    ws_hvulns = wb.create_sheet("Host Vulnerabilities")
    build_vulns_sheet(ws_hvulns, host_vulns, "host")

    # Sheet 4 — Host Packages
    ws_pkgs = wb.create_sheet("Host Packages")
    build_host_packages(ws_pkgs, host_pkgs)

    # Sheet 5 — Container Inventory
    ws_containers = wb.create_sheet("Container Inventory")
    build_container_inventory(ws_containers, container_data)

    # Sheet 6 — Container Vulnerabilities (all images combined)
    all_container_vulns = [v for _, _, cvulns in container_data for v in cvulns]
    ws_cvulns = wb.create_sheet("Container Vulnerabilities")
    build_vulns_sheet(ws_cvulns, all_container_vulns, "containers")

    # Per-container sheets (if ≤10 containers, else skip to keep workbook manageable)
    if len(container_data) <= 10:
        for img_name, pkgs, vulns in container_data:
            safe = img_name.replace("/", "_").replace(":", "_")[:28]
            ws_ci = wb.create_sheet(f"C {safe}")
            build_vulns_sheet(ws_ci, vulns, img_name)

    wb.save(OUTPUT_FILE)
    print(f"\n[✓] Report saved: {OUTPUT_FILE}")
    print(f"    Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
